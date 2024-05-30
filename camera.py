import os
import cv2
from datetime import datetime
from openpyxl import load_workbook, Workbook
from flask import redirect

filename_to_name = {
    'David.jpg': 'David',
    'Elon.jpg': 'Elon',
    'Emily.jpg': 'Emily'
}

reference_image_paths = list(filename_to_name.keys())
reference_images = [cv2.imread(path) for path in reference_image_paths]
reference_grays = [cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) for img in reference_images]

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

def write_to_excel(name):
    wb = load_workbook('entry_log.xlsx') if os.path.exists('entry_log.xlsx') else Workbook()
    ws = wb.active
    ws.append([name, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save('entry_log.xlsx')

class Video(object):
    
    def __init__(self):
        self.video = cv2.VideoCapture(0)
    
    def __del__(self):
        self.video.release()

    def get_frame(self):
        authenticated = False
        ret, frame = self.video.read()
       
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(30, 30))
        
        for (x, y, w, h) in faces:
            roi_gray = gray[y:y + h, x:x + w]
            
            
           
            for reference_gray, filename in zip(reference_grays, reference_image_paths):
                hist_ref = cv2.calcHist([reference_gray], [0], None, [256], [0, 256])
                hist_roi = cv2.calcHist([roi_gray], [0], None, [256], [0, 256])
                similarity = cv2.compareHist(hist_ref, hist_roi, cv2.HISTCMP_CORREL)
                if similarity > 0.8:
                    name = filename_to_name[filename]
                    authenticated = True
                    write_to_excel(name)
                
                    break
            if authenticated:
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, "Authenticated", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                
            else:
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.putText(frame, "Not Authenticated", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                
        ret, jpg = cv2.imencode('.jpg', frame)
        return jpg.tobytes(), authenticated
    